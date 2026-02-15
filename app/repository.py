from __future__ import annotations

import shutil
import uuid
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any, Callable

from filelock import FileLock
from openpyxl import Workbook, load_workbook

from app.config import settings
from app.constants import (
    ABSENCES_HEADERS,
    DESKS_HEADERS,
    META_HEADERS,
    RESERVATIONS_HEADERS,
    USERS_HEADERS,
)
from app.domain import normalize_bool
from app.models import AbsenceRecord, DeskRecord, ReservationRecord, UserRecord


@dataclass
class Tables:
    users: list[dict[str, Any]]
    desks: list[dict[str, Any]]
    reservations: list[dict[str, Any]]
    absences: list[dict[str, Any]]
    meta: list[dict[str, Any]]


class ExcelRepository:
    def __init__(self) -> None:
        self.data_file = settings.data_file
        self.backup_dir = settings.backup_dir
        self.lock = FileLock(str(settings.lock_file))

    def init_storage(self) -> None:
        self.data_file.parent.mkdir(parents=True, exist_ok=True)
        self.backup_dir.mkdir(parents=True, exist_ok=True)
        settings.lock_file.parent.mkdir(parents=True, exist_ok=True)
        if self.data_file.exists():
            return

        wb = Workbook()
        default = wb.active
        wb.remove(default)
        for sheet_name, headers in self._sheet_headers().items():
            ws = wb.create_sheet(sheet_name)
            ws.append(headers)
        wb.save(self.data_file)

    def list_users(self) -> list[UserRecord]:
        tables = self._read_tables()
        return [
            UserRecord(
                user_id=row["user_id"],
                name=self._normalize_user_name(row),
                email=row.get("email") or None,
                enabled=normalize_bool(row["enabled"]),
                is_admin=normalize_bool(row["is_admin"]),
                created_at=self._parse_datetime(row["created_at"]),
            )
            for row in tables.users
            if row.get("user_id")
        ]

    def get_user_by_name(self, name: str) -> UserRecord | None:
        normalized_name = name.strip().lower()
        for user in self.list_users():
            if user.name.strip().lower() == normalized_name:
                return user
        return None

    def get_user_by_email(self, email: str) -> UserRecord | None:
        normalized = email.lower()
        for user in self.list_users():
            if user.email and user.email.lower() == normalized:
                return user
        return None

    def get_user(self, user_id: str) -> UserRecord | None:
        for user in self.list_users():
            if user.user_id == user_id:
                return user
        return None

    def upsert_user(
        self,
        name: str,
        enabled: bool = True,
        is_admin: bool = False,
        email: str | None = None,
    ) -> UserRecord:
        now = datetime.utcnow().isoformat()
        normalized_name = name.strip()
        normalized_email = email.lower().strip() if email else None

        def mutate(tables: Tables) -> dict[str, Any]:
            for row in tables.users:
                row_name = self._normalize_user_name(row)
                if row_name.strip().lower() == normalized_name.lower():
                    row["name"] = normalized_name
                    if normalized_email is not None:
                        row["email"] = normalized_email
                    row["enabled"] = enabled
                    row["is_admin"] = is_admin
                    return row
            row = {
                "user_id": uuid.uuid4().hex,
                "name": normalized_name,
                "email": normalized_email,
                "enabled": enabled,
                "is_admin": is_admin,
                "created_at": now,
            }
            tables.users.append(row)
            return row

        row = self._write_tables(mutate)
        return UserRecord(
            user_id=row["user_id"],
            name=self._normalize_user_name(row),
            email=row.get("email") or None,
            enabled=normalize_bool(row["enabled"]),
            is_admin=normalize_bool(row["is_admin"]),
            created_at=self._parse_datetime(row["created_at"]),
        )

    def list_desks(self) -> list[DeskRecord]:
        tables = self._read_tables()
        return [
            DeskRecord(
                desk_id=row["desk_id"],
                label=row["label"],
                enabled=normalize_bool(row["enabled"]),
                owner_user_id=row.get("owner_user_id") or None,
            )
            for row in tables.desks
            if row.get("desk_id")
        ]

    def upsert_desk(
        self,
        label: str,
        enabled: bool = True,
        owner_user_id: str | None = None,
        desk_id: str | None = None,
    ) -> DeskRecord:
        def mutate(tables: Tables) -> dict[str, Any]:
            target_id = desk_id
            if target_id:
                for row in tables.desks:
                    if row.get("desk_id") == target_id:
                        row["label"] = label
                        row["enabled"] = enabled
                        row["owner_user_id"] = owner_user_id
                        return row

            new_row = {
                "desk_id": target_id or uuid.uuid4().hex,
                "label": label,
                "enabled": enabled,
                "owner_user_id": owner_user_id,
            }
            tables.desks.append(new_row)
            return new_row

        row = self._write_tables(mutate)
        return DeskRecord(
            desk_id=row["desk_id"],
            label=row["label"],
            enabled=normalize_bool(row["enabled"]),
            owner_user_id=row.get("owner_user_id") or None,
        )

    def get_desk(self, desk_id: str) -> DeskRecord | None:
        for desk in self.list_desks():
            if desk.desk_id == desk_id:
                return desk
        return None

    def list_reservations(
        self,
        start_date: date | None = None,
        end_date: date | None = None,
    ) -> list[ReservationRecord]:
        tables = self._read_tables()
        rows: list[ReservationRecord] = []
        for row in tables.reservations:
            if not row.get("reservation_id"):
                continue
            value_date = self._parse_date(row["date"])
            if start_date and value_date < start_date:
                continue
            if end_date and value_date > end_date:
                continue
            rows.append(
                ReservationRecord(
                    reservation_id=row["reservation_id"],
                    user_id=row["user_id"],
                    desk_id=row["desk_id"],
                    date=value_date,
                    slot=row["slot"],
                    created_at=self._parse_datetime(row["created_at"]),
                    updated_at=self._parse_datetime(row["updated_at"]),
                    auto=False,
                )
            )
        return rows

    def get_reservation(self, reservation_id: str) -> ReservationRecord | None:
        for item in self.list_reservations():
            if item.reservation_id == reservation_id:
                return item
        return None

    def create_reservation(self, user_id: str, desk_id: str, value_date: date, slot: str) -> ReservationRecord:
        now = datetime.utcnow().isoformat()

        def mutate(tables: Tables) -> dict[str, Any]:
            for existing in tables.reservations:
                if (
                    self._parse_date(existing["date"]) == value_date
                    and existing["slot"] == slot
                    and existing["desk_id"] == desk_id
                ):
                    raise ValueError("Desk already reserved")
                if (
                    self._parse_date(existing["date"]) == value_date
                    and existing["slot"] == slot
                    and existing["user_id"] == user_id
                ):
                    raise ValueError("User already has a desk in this slot")
            row = {
                "reservation_id": uuid.uuid4().hex,
                "user_id": user_id,
                "desk_id": desk_id,
                "date": value_date.isoformat(),
                "slot": slot,
                "created_at": now,
                "updated_at": now,
            }
            tables.reservations.append(row)
            return row

        try:
            row = self._write_tables(mutate)
        except ValueError as exc:
            raise ValueError(str(exc)) from exc
        return ReservationRecord(
            reservation_id=row["reservation_id"],
            user_id=row["user_id"],
            desk_id=row["desk_id"],
            date=self._parse_date(row["date"]),
            slot=row["slot"],
            created_at=self._parse_datetime(row["created_at"]),
            updated_at=self._parse_datetime(row["updated_at"]),
            auto=False,
        )

    def update_reservation(
        self,
        reservation_id: str,
        user_id: str,
        desk_id: str,
        value_date: date,
        slot: str,
    ) -> ReservationRecord | None:
        now = datetime.utcnow().isoformat()

        def mutate(tables: Tables) -> dict[str, Any] | None:
            for existing in tables.reservations:
                if existing.get("reservation_id") == reservation_id:
                    continue
                if (
                    self._parse_date(existing["date"]) == value_date
                    and existing["slot"] == slot
                    and existing["desk_id"] == desk_id
                ):
                    raise ValueError("Desk already reserved")
                if (
                    self._parse_date(existing["date"]) == value_date
                    and existing["slot"] == slot
                    and existing["user_id"] == user_id
                ):
                    raise ValueError("User already has a desk in this slot")
            for row in tables.reservations:
                if row.get("reservation_id") == reservation_id:
                    row["user_id"] = user_id
                    row["desk_id"] = desk_id
                    row["date"] = value_date.isoformat()
                    row["slot"] = slot
                    row["updated_at"] = now
                    return row
            return None

        try:
            row = self._write_tables(mutate)
        except ValueError as exc:
            raise ValueError(str(exc)) from exc
        if not row:
            return None
        return ReservationRecord(
            reservation_id=row["reservation_id"],
            user_id=row["user_id"],
            desk_id=row["desk_id"],
            date=self._parse_date(row["date"]),
            slot=row["slot"],
            created_at=self._parse_datetime(row["created_at"]),
            updated_at=self._parse_datetime(row["updated_at"]),
            auto=False,
        )

    def delete_reservation(self, reservation_id: str) -> bool:
        def mutate(tables: Tables) -> bool:
            initial = len(tables.reservations)
            tables.reservations = [
                row for row in tables.reservations if row.get("reservation_id") != reservation_id
            ]
            return len(tables.reservations) != initial

        return bool(self._write_tables(mutate))

    def list_absences(self) -> list[AbsenceRecord]:
        tables = self._read_tables()
        rows: list[AbsenceRecord] = []
        for row in tables.absences:
            if not row.get("absence_id"):
                continue
            rows.append(
                AbsenceRecord(
                    absence_id=row["absence_id"],
                    owner_user_id=row["owner_user_id"],
                    desk_id=row["desk_id"],
                    date=self._parse_date(row["date"]),
                    slot=row["slot"],
                    created_at=self._parse_datetime(row["created_at"]),
                )
            )
        return rows

    def upsert_absence(
        self,
        owner_user_id: str,
        desk_id: str,
        value_date: date,
        slot: str,
        released: bool,
    ) -> None:
        def mutate(tables: Tables) -> None:
            matches = [
                row
                for row in tables.absences
                if row.get("owner_user_id") == owner_user_id
                and row.get("desk_id") == desk_id
                and self._parse_date(str(row.get("date"))) == value_date
                and row.get("slot") == slot
            ]
            if released and not matches:
                tables.absences.append(
                    {
                        "absence_id": uuid.uuid4().hex,
                        "owner_user_id": owner_user_id,
                        "desk_id": desk_id,
                        "date": value_date.isoformat(),
                        "slot": slot,
                        "created_at": datetime.utcnow().isoformat(),
                    }
                )
            if not released and matches:
                ids = {row.get("absence_id") for row in matches}
                tables.absences = [
                    row for row in tables.absences if row.get("absence_id") not in ids
                ]

        self._write_tables(mutate)

    def stats(self) -> dict[str, int]:
        users = self.list_users()
        desks = self.list_desks()
        reservations = self.list_reservations()
        return {
            "total_reservations": len(reservations),
            "active_users": len([u for u in users if u.enabled]),
            "enabled_desks": len([d for d in desks if d.enabled]),
        }

    def _sheet_headers(self) -> dict[str, list[str]]:
        return {
            "users": USERS_HEADERS,
            "desks": DESKS_HEADERS,
            "reservations": RESERVATIONS_HEADERS,
            "absences": ABSENCES_HEADERS,
            "meta": META_HEADERS,
        }

    def _read_tables(self) -> Tables:
        self.init_storage()
        wb = load_workbook(self.data_file)
        try:
            return Tables(
                users=self._read_sheet(wb, "users", USERS_HEADERS),
                desks=self._read_sheet(wb, "desks", DESKS_HEADERS),
                reservations=self._read_sheet(wb, "reservations", RESERVATIONS_HEADERS),
                absences=self._read_sheet(wb, "absences", ABSENCES_HEADERS),
                meta=self._read_sheet(wb, "meta", META_HEADERS),
            )
        finally:
            wb.close()

    def _write_tables(self, mutator: Callable[[Tables], Any]) -> Any:
        self.init_storage()
        with self.lock:
            wb = load_workbook(self.data_file)
            try:
                tables = Tables(
                    users=self._read_sheet(wb, "users", USERS_HEADERS),
                    desks=self._read_sheet(wb, "desks", DESKS_HEADERS),
                    reservations=self._read_sheet(wb, "reservations", RESERVATIONS_HEADERS),
                    absences=self._read_sheet(wb, "absences", ABSENCES_HEADERS),
                    meta=self._read_sheet(wb, "meta", META_HEADERS),
                )
                result = mutator(tables)
                self._write_sheet(wb, "users", USERS_HEADERS, tables.users)
                self._write_sheet(wb, "desks", DESKS_HEADERS, tables.desks)
                self._write_sheet(wb, "reservations", RESERVATIONS_HEADERS, tables.reservations)
                self._write_sheet(wb, "absences", ABSENCES_HEADERS, tables.absences)
                self._write_sheet(wb, "meta", META_HEADERS, tables.meta)
                self._persist_workbook(wb)
                return result
            finally:
                wb.close()

    def _persist_workbook(self, workbook: Workbook) -> None:
        with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            temp_path = Path(tmp.name)
        try:
            workbook.save(temp_path)
            if self.data_file.exists():
                stamp = datetime.utcnow().strftime("%Y%m%d%H%M%S%f")
                backup_path = self.backup_dir / f"reservations-{stamp}.xlsx"
                shutil.copy2(self.data_file, backup_path)
            temp_path.replace(self.data_file)
        finally:
            if temp_path.exists():
                temp_path.unlink(missing_ok=True)

    def _read_sheet(self, workbook: Workbook, name: str, headers: list[str]) -> list[dict[str, Any]]:
        ws = workbook[name]
        rows: list[dict[str, Any]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(item is None for item in row):
                continue
            payload: dict[str, Any] = {}
            for index, header in enumerate(headers):
                payload[header] = row[index] if index < len(row) else None
            rows.append(payload)
        return rows

    def _write_sheet(
        self,
        workbook: Workbook,
        name: str,
        headers: list[str],
        rows: list[dict[str, Any]],
    ) -> None:
        if name not in workbook.sheetnames:
            workbook.create_sheet(name)
        ws = workbook[name]
        ws.delete_rows(1, ws.max_row)
        ws.append(headers)
        for row in rows:
            ws.append([row.get(header) for header in headers])

    def _parse_date(self, raw: Any) -> date:
        if isinstance(raw, date) and not isinstance(raw, datetime):
            return raw
        if isinstance(raw, datetime):
            return raw.date()
        return date.fromisoformat(str(raw))

    def _parse_datetime(self, raw: Any) -> datetime:
        if isinstance(raw, datetime):
            return raw
        return datetime.fromisoformat(str(raw))

    def _normalize_user_name(self, row: dict[str, Any]) -> str:
        name = str(row.get("name") or "").strip()
        if name:
            return name
        email = str(row.get("email") or "").strip()
        if "@" in email:
            return email.split("@", 1)[0]
        return email or "user"
